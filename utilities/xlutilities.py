import openpyxl
def getrow(filename,sheetname):
    wb=openpyxl.load_workbook(filename)
    ws=wb[sheetname]
    return ws.max_row
def readrow(filename,sheetname,rownum,columnnum):
    wb=openpyxl.load_workbook(filename)
    ws=wb[sheetname]
    return ws.cell(rownum,columnnum).value
